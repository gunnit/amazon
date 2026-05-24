"""Tests for the strategic recommendations XLSX export builder.

The builder runs from the API layer in exports.py with already-loaded
SQLAlchemy rows. The test exercises it directly with lightweight stand-
in objects so it does not need a live database, only openpyxl.
"""
from __future__ import annotations

import io
from types import SimpleNamespace
from datetime import datetime
from uuid import uuid4

from openpyxl import load_workbook

from app.services.strategic_recommendations_export import (
    build_recommendations_workbook_bytes,
)


def _make_rec(
    *,
    category: str = "pricing",
    priority: str = "high",
    title: str = "Lower price on B0TEST",
    rationale: str = "Sales are flat and ACoS is rising.",
    expected_impact: str = "Reduce ACoS by 4pp over next 14 days.",
    status: str = "pending",
    asin: str = "B0TESTASIN",
    account_id: str | None = None,
):
    return SimpleNamespace(
        id=uuid4(),
        category=category,
        priority=priority,
        title=title,
        rationale=rationale,
        expected_impact=expected_impact,
        status=status,
        generated_at=datetime(2026, 5, 24, 10, 30),
        context={"asins": [asin]} if asin else None,
        account_id=account_id,
    )


def test_workbook_contains_summary_and_details_sheets_with_scope():
    account_id = "11111111-1111-1111-1111-111111111111"
    account_names = {account_id: "Test Seller IT"}
    recs = [
        _make_rec(category="pricing", priority="high"),
        _make_rec(category="advertising", priority="medium", asin="B0TESTASIN"),
    ]

    payload = build_recommendations_workbook_bytes(
        recs,
        account_names=account_names,
        language="en",
        scope_account_id=account_id,
        scope_asin="B0TESTASIN",
    )

    wb = load_workbook(io.BytesIO(payload))
    sheet_names = set(wb.sheetnames)
    assert "Summary" in sheet_names
    assert "Recommendations" in sheet_names

    summary = wb["Summary"]
    # Banner shows the scoped account name + ASIN
    cells = [cell.value for cell in summary["B"][:8] if cell.value is not None]
    assert "Test Seller IT" in cells
    assert "B0TESTASIN" in cells

    details = wb["Recommendations"]
    headers = [cell.value for cell in details[1]]
    assert "Title" in headers
    assert "Rationale" in headers
    # Both recs land in the table
    assert details.max_row >= 3  # header + 2 rows


def test_workbook_renders_in_italian_locale():
    payload = build_recommendations_workbook_bytes(
        [_make_rec(category="content", priority="low")],
        account_names={},
        language="it",
    )
    wb = load_workbook(io.BytesIO(payload))
    summary = wb["Riepilogo"]
    headers_seen = {cell.value for row in summary.iter_rows() for cell in row}
    # Italian summary banner title
    assert any("Report Raccomandazioni" in str(v) for v in headers_seen if v)
    # Category label is localized
    details = wb["Raccomandazioni"]
    body = {cell.value for row in details.iter_rows() for cell in row}
    assert "Contenuti" in body
