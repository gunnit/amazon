"""Catalog write operations against the Amazon SP-API Listings Items API.

Reads from the local Product table, pushes changes through SP-API,
then mirrors the successful change locally and records an audit row.
"""
from __future__ import annotations

import io
import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Sequence
from uuid import UUID

import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AmazonAPIError
from app.models.amazon_account import AmazonAccount, AccountType
from app.models.catalog_change_log import CatalogChangeLog
from app.models.product import Product
from app.schemas.catalog import (
    ASIN_PATTERN,
    BulkErrorCode,
    BulkListingUpdateResult,
    BulkResult,
    BulkRowError,
    CatalogChangeField,
    CatalogChangeStatus,
    ImportResult,
    PriceUpdate,
    PriceUpdateResult,
    ProductImportRow,
    AvailabilityResult,
)

logger = logging.getLogger(__name__)

_ASIN_RE = re.compile(ASIN_PATTERN)
IMPORT_COLUMNS = ("asin", "sku", "title", "brand", "category")

LISTING_TEXT_FIELDS = {
    "title": "item_name",
    "bullet_1": "bullet_point",
    "bullet_2": "bullet_point",
    "bullet_3": "bullet_point",
    "bullet_4": "bullet_point",
    "bullet_5": "bullet_point",
    "description": "product_description",
    "search_terms": "generic_keyword",
}


class CatalogOperationError(Exception):
    """Raised when a catalog write operation cannot be completed."""


class CatalogService:
    """Catalog management service backed by SP-API Listings."""

    def __init__(self, db: AsyncSession, user_id: Optional[UUID] = None):
        self.db = db
        self.user_id = user_id

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    async def _load_organization(self, account: AmazonAccount):
        from app.models.user import Organization
        result = await self.db.execute(
            select(Organization).where(Organization.id == account.organization_id)
        )
        return result.scalar_one_or_none()

    def _create_sp_api_client(self, account: AmazonAccount, organization=None):
        from app.core.amazon.credentials import resolve_credentials
        from app.core.amazon.sp_api_client import SPAPIClient, resolve_marketplace

        credentials = resolve_credentials(account, organization)
        marketplace = resolve_marketplace(account.marketplace_country)
        return SPAPIClient(credentials, marketplace, account_type=account.account_type.value)

    async def _require_seller_account(self, account_id: UUID) -> AmazonAccount:
        result = await self.db.execute(
            select(AmazonAccount).where(AmazonAccount.id == account_id)
        )
        account = result.scalar_one_or_none()
        if not account:
            raise CatalogOperationError(f"Account {account_id} not found")
        if account.account_type != AccountType.SELLER:
            raise CatalogOperationError(
                "Listing writes are only available for Seller Central accounts"
            )
        if not account.seller_id:
            raise CatalogOperationError(
                f"Account {account.account_name} is missing a Seller ID"
            )
        return account

    async def _load_product(self, account_id: UUID, asin: str) -> Optional[Product]:
        result = await self.db.execute(
            select(Product).where(
                Product.account_id == account_id,
                Product.asin == asin,
            )
        )
        return result.scalar_one_or_none()

    def _audit(
        self,
        *,
        organization_id: UUID,
        account_id: UUID,
        asin: Optional[str],
        sku: Optional[str],
        field: CatalogChangeField,
        old_value: Any,
        new_value: Any,
        status: CatalogChangeStatus,
        sp_api_error: Optional[str] = None,
    ) -> None:
        entry = CatalogChangeLog(
            organization_id=organization_id,
            account_id=account_id,
            user_id=self.user_id,
            asin=asin,
            sku=sku,
            field=field.value,
            old_value=old_value,
            new_value=new_value,
            sp_api_status=status.value,
            sp_api_error=sp_api_error,
        )
        self.db.add(entry)

    # ------------------------------------------------------------------
    # Bulk listing updates (title / bullets / description / keywords)
    # ------------------------------------------------------------------

    async def bulk_update_from_excel(
        self,
        account_id: UUID,
        file_bytes: bytes,
        product_type: str = "PRODUCT",
    ) -> BulkResult[BulkListingUpdateResult]:
        """Apply bulk listing content updates from an Excel file.

        Expected columns (all optional except sku): sku, title, bullet_1..5,
        description, search_terms.
        """
        account = await self._require_seller_account(account_id)
        organization = await self._load_organization(account)
        client = self._create_sp_api_client(account, organization)

        try:
            df = pd.read_excel(io.BytesIO(file_bytes))
        except Exception as exc:  # noqa: BLE001
            raise CatalogOperationError(f"Could not parse Excel file: {exc}") from exc

        if "sku" not in df.columns:
            raise CatalogOperationError("Excel must contain a 'sku' column")

        successes: List[BulkListingUpdateResult] = []
        errors: List[BulkRowError] = []
        skipped = 0

        for row_idx, row in df.iterrows():
            row_number = int(row_idx) + 2  # +1 for 0-based, +1 for header
            sku = str(row.get("sku") or "").strip()
            if not sku:
                skipped += 1
                continue

            attributes = _row_to_listing_attributes(row)
            if not attributes:
                skipped += 1
                continue

            old_snapshot = await self._product_snapshot_by_sku(account_id, sku)
            new_snapshot = _row_listing_snapshot(row)

            try:
                client.update_listing_attributes(
                    seller_id=account.seller_id,
                    sku=sku,
                    product_type=product_type,
                    attributes=attributes,
                )
                await self._mirror_local_listing(account_id, sku, row)
                successes.append(
                    BulkListingUpdateResult(sku=sku, fields=list(attributes.keys()))
                )
                self._audit(
                    organization_id=account.organization_id,
                    account_id=account_id,
                    asin=old_snapshot.get("asin") if old_snapshot else None,
                    sku=sku,
                    field=CatalogChangeField.LISTING,
                    old_value=old_snapshot,
                    new_value=new_snapshot,
                    status=CatalogChangeStatus.SUCCESS,
                )
            except AmazonAPIError as exc:
                logger.warning("SP-API listing update failed for %s: %s", sku, exc)
                errors.append(
                    BulkRowError(
                        row=row_number,
                        sku=sku,
                        error=str(exc),
                        code=BulkErrorCode.SP_API_ERROR,
                    )
                )
                self._audit(
                    organization_id=account.organization_id,
                    account_id=account_id,
                    asin=old_snapshot.get("asin") if old_snapshot else None,
                    sku=sku,
                    field=CatalogChangeField.LISTING,
                    old_value=old_snapshot,
                    new_value=new_snapshot,
                    status=CatalogChangeStatus.FAILED,
                    sp_api_error=str(exc),
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("Unexpected failure updating listing %s", sku)
                errors.append(
                    BulkRowError(
                        row=row_number,
                        sku=sku,
                        error=str(exc),
                        code=BulkErrorCode.UNEXPECTED_ERROR,
                    )
                )

        await self.db.flush()

        return BulkResult[BulkListingUpdateResult](
            account_id=account_id,
            total=int(len(df)),
            succeeded=len(successes),
            failed=len(errors),
            skipped=skipped,
            successes=successes,
            errors=errors,
        )

    async def _mirror_local_listing(self, account_id: UUID, sku: str, row: pd.Series) -> None:
        """Reflect the update in the local catalog cache so the UI shows it immediately."""
        result = await self.db.execute(
            select(Product).where(Product.account_id == account_id, Product.sku == sku)
        )
        product = result.scalar_one_or_none()
        if not product:
            return

        if "title" in row and pd.notna(row.get("title")):
            product.title = str(row["title"])
        if "brand" in row and pd.notna(row.get("brand")):
            product.brand = str(row["brand"])
        if "category" in row and pd.notna(row.get("category")):
            product.category = str(row["category"])

    async def _product_snapshot_by_sku(
        self, account_id: UUID, sku: str
    ) -> Optional[Dict[str, Any]]:
        result = await self.db.execute(
            select(Product).where(Product.account_id == account_id, Product.sku == sku)
        )
        product = result.scalar_one_or_none()
        if not product:
            return None
        return {
            "asin": product.asin,
            "sku": product.sku,
            "title": product.title,
            "brand": product.brand,
            "category": product.category,
        }

    # ------------------------------------------------------------------
    # Manual catalog import (CSV / Excel)
    # ------------------------------------------------------------------

    async def import_products_from_file(
        self,
        account_id: UUID,
        file_bytes: bytes,
        filename: str,
    ) -> ImportResult:
        """Create or update local Product rows from a CSV/Excel upload.

        Works for both Seller and Vendor accounts: rows are written straight to
        the local catalog and never pushed to SP-API. Existing synced metadata is
        only overwritten when the incoming cell is non-empty.
        """
        rows, errors = parse_import_rows(file_bytes, filename)

        successes: List[ProductImportRow] = []
        for parsed in rows:
            asin = parsed["asin"]
            existing = await self._load_product(account_id, asin)
            created = existing is None

            if existing is None:
                existing = Product(account_id=account_id, asin=asin, is_active=True)
                self.db.add(existing)

            if parsed.get("sku"):
                existing.sku = parsed["sku"]
            if parsed.get("title"):
                existing.title = parsed["title"]
            if parsed.get("brand"):
                existing.brand = parsed["brand"]
            if parsed.get("category"):
                existing.category = parsed["category"]
            existing.is_active = True
            existing.source = "manual_import"

            successes.append(
                ProductImportRow(
                    asin=asin,
                    sku=existing.sku,
                    title=existing.title,
                    brand=existing.brand,
                    category=existing.category,
                    created=created,
                )
            )

        await self.db.flush()

        total = len(rows) + len(errors)
        return ImportResult(
            account_id=account_id,
            total=total,
            succeeded=len(successes),
            failed=len(errors),
            skipped=0,
            successes=successes,
            errors=errors,
        )

    # ------------------------------------------------------------------
    # Price management
    # ------------------------------------------------------------------

    async def update_prices_bulk(
        self,
        account_id: UUID,
        updates: Sequence[PriceUpdate],
        product_type: str = "PRODUCT",
    ) -> BulkResult[PriceUpdateResult]:
        """Push a list of price updates to SP-API."""
        account = await self._require_seller_account(account_id)
        organization = await self._load_organization(account)
        client = self._create_sp_api_client(account, organization)

        currency = _marketplace_currency(account.marketplace_country)
        successes: List[PriceUpdateResult] = []
        errors: List[BulkRowError] = []

        for entry in updates:
            asin = entry.asin
            sku = entry.sku
            new_price = entry.price

            product = await self._resolve_product(account_id, asin=asin, sku=sku)
            if not product:
                errors.append(
                    BulkRowError(
                        asin=asin,
                        sku=sku,
                        error="Product not found for this account",
                        code=BulkErrorCode.PRODUCT_NOT_FOUND,
                    )
                )
                continue
            if not product.sku:
                errors.append(
                    BulkRowError(
                        asin=product.asin,
                        sku=sku,
                        error="Product has no SKU; cannot call SP-API",
                        code=BulkErrorCode.MISSING_SKU,
                    )
                )
                continue

            old_price = product.current_price

            try:
                client.update_listing_price(
                    seller_id=account.seller_id,
                    sku=product.sku,
                    product_type=product_type,
                    price=new_price,
                    currency=currency,
                )
                product.current_price = new_price
                successes.append(
                    PriceUpdateResult(asin=product.asin, sku=product.sku, price=new_price)
                )
                self._audit(
                    organization_id=account.organization_id,
                    account_id=account_id,
                    asin=product.asin,
                    sku=product.sku,
                    field=CatalogChangeField.PRICE,
                    old_value={"price": str(old_price) if old_price is not None else None, "currency": currency},
                    new_value={"price": str(new_price), "currency": currency},
                    status=CatalogChangeStatus.SUCCESS,
                )
            except AmazonAPIError as exc:
                errors.append(
                    BulkRowError(
                        asin=product.asin,
                        sku=product.sku,
                        error=str(exc),
                        code=BulkErrorCode.SP_API_ERROR,
                    )
                )
                self._audit(
                    organization_id=account.organization_id,
                    account_id=account_id,
                    asin=product.asin,
                    sku=product.sku,
                    field=CatalogChangeField.PRICE,
                    old_value={"price": str(old_price) if old_price is not None else None, "currency": currency},
                    new_value={"price": str(new_price), "currency": currency},
                    status=CatalogChangeStatus.FAILED,
                    sp_api_error=str(exc),
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("Unexpected failure updating price for %s", product.sku)
                errors.append(
                    BulkRowError(
                        asin=product.asin,
                        sku=product.sku,
                        error=str(exc),
                        code=BulkErrorCode.UNEXPECTED_ERROR,
                    )
                )

        await self.db.flush()

        return BulkResult[PriceUpdateResult](
            account_id=account_id,
            total=len(updates),
            succeeded=len(successes),
            failed=len(errors),
            successes=successes,
            errors=errors,
        )

    # ------------------------------------------------------------------
    # Availability toggle
    # ------------------------------------------------------------------

    async def toggle_availability(
        self,
        account_id: UUID,
        asin: str,
        is_available: bool,
        quantity: Optional[int] = None,
        product_type: str = "PRODUCT",
    ) -> AvailabilityResult:
        """Enable or disable a product by setting fulfillment quantity."""
        account = await self._require_seller_account(account_id)
        organization = await self._load_organization(account)
        client = self._create_sp_api_client(account, organization)

        product = await self._load_product(account_id, asin)
        if not product:
            raise CatalogOperationError(f"Product {asin} not found for this account")
        if not product.sku:
            raise CatalogOperationError(f"Product {asin} has no SKU")

        effective_quantity = 0 if not is_available else max(1, int(quantity or 1))
        old_value = {
            "is_available": bool(product.is_available),
            "is_active": bool(product.is_active),
        }
        new_value = {
            "is_available": is_available,
            "is_active": bool(is_available),
            "pushed_quantity": effective_quantity,
        }

        try:
            client.set_listing_quantity(
                seller_id=account.seller_id,
                sku=product.sku,
                product_type=product_type,
                quantity=effective_quantity,
            )
        except AmazonAPIError as exc:
            self._audit(
                organization_id=account.organization_id,
                account_id=account_id,
                asin=product.asin,
                sku=product.sku,
                field=CatalogChangeField.AVAILABILITY,
                old_value=old_value,
                new_value=new_value,
                status=CatalogChangeStatus.FAILED,
                sp_api_error=str(exc),
            )
            await self.db.flush()
            raise CatalogOperationError(str(exc)) from exc

        product.is_available = is_available
        product.is_active = bool(is_available)

        self._audit(
            organization_id=account.organization_id,
            account_id=account_id,
            asin=product.asin,
            sku=product.sku,
            field=CatalogChangeField.AVAILABILITY,
            old_value=old_value,
            new_value=new_value,
            status=CatalogChangeStatus.SUCCESS,
        )
        await self.db.flush()

        return AvailabilityResult(
            asin=product.asin,
            sku=product.sku,
            is_available=is_available,
            pushed_quantity=effective_quantity,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    async def _resolve_product(
        self,
        account_id: UUID,
        asin: Optional[str] = None,
        sku: Optional[str] = None,
    ) -> Optional[Product]:
        if not asin and not sku:
            return None
        stmt = select(Product).where(Product.account_id == account_id)
        if asin:
            stmt = stmt.where(Product.asin == asin)
        if sku:
            stmt = stmt.where(Product.sku == sku)
        result = await self.db.execute(stmt)
        return result.scalars().first()

    @staticmethod
    def generate_template_bytes() -> bytes:
        """Return a styled Excel template for bulk listing updates.

        The ``listings`` sheet (first sheet, parsed on re-upload) carries only
        the header row in the exact order the importer expects. A sample row and
        per-column guidance live on a separate ``Instructions`` sheet so they are
        never picked up as a real update.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        columns = ["sku", "asin", "title", "bullet_1", "bullet_2", "bullet_3",
                   "bullet_4", "bullet_5", "description", "search_terms"]
        widths = {
            "sku": 20, "asin": 16, "title": 50,
            "bullet_1": 40, "bullet_2": 40, "bullet_3": 40,
            "bullet_4": 40, "bullet_5": 40,
            "description": 60, "search_terms": 45,
        }

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "listings"
        ws.append(columns)
        for idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(idx)].width = widths[column]
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        guide = [
            ("sku", "Required", "Seller SKU of the product to update. Rows without a SKU are ignored."),
            ("asin", "Optional", "ASIN, for your reference only. Not modified by the import."),
            ("title", "Optional", "Listing title. Max 200 characters."),
            ("bullet_1", "Optional", "First bullet point. Max 255 characters each."),
            ("bullet_2", "Optional", "Second bullet point."),
            ("bullet_3", "Optional", "Third bullet point."),
            ("bullet_4", "Optional", "Fourth bullet point."),
            ("bullet_5", "Optional", "Fifth bullet point."),
            ("description", "Optional", "Product description. Max 2000 characters."),
            ("search_terms", "Optional", "Backend keywords, space-separated. Max 250 bytes."),
        ]
        example = {
            "sku": "EXAMPLE-SKU-001",
            "asin": "B08XXXXXXX",
            "title": "Stainless Steel Water Bottle 750ml",
            "bullet_1": "Keeps drinks cold for 24h and hot for 12h",
            "bullet_2": "Leak-proof screw cap with carry loop",
            "bullet_3": "BPA-free food-grade stainless steel",
            "bullet_4": "Fits standard car cup holders",
            "bullet_5": "Dishwasher safe",
            "description": "Double-walled vacuum insulated bottle for everyday use.",
            "search_terms": "water bottle insulated thermos flask reusable",
        }

        info = wb.create_sheet("Instructions")
        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wrap = Alignment(vertical="top", wrap_text=True)

        info["A1"] = "Bulk Listing Update — Template Guide"
        info["A1"].font = Font(bold=True, size=14, color="1F4E79")
        info["A3"] = (
            "Fill in the 'listings' sheet, one product per row. Only the SKU is "
            "required; leave any field blank to keep its current value. Do not "
            "rename the sheet or its column headers."
        )
        info["A3"].alignment = wrap
        info.merge_cells("A3:C3")
        info.row_dimensions[3].height = 45

        head_row = 5
        for idx, label in enumerate(("Column", "Required", "Description"), start=1):
            cell = info.cell(row=head_row, column=idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        for offset, (name, requirement, desc) in enumerate(guide, start=1):
            row = head_row + offset
            for idx, value in enumerate((name, requirement, desc), start=1):
                cell = info.cell(row=row, column=idx, value=value)
                cell.border = border
                cell.alignment = wrap

        example_title = head_row + len(guide) + 2
        info.cell(row=example_title, column=1, value="Example row (for reference only — do NOT paste into 'listings' as-is):")
        info.cell(row=example_title, column=1).font = Font(bold=True, italic=True, color="7A7A7A")
        info.merge_cells(start_row=example_title, start_column=1, end_row=example_title, end_column=3)

        ex_head = example_title + 1
        for idx, column in enumerate(columns, start=1):
            cell = info.cell(row=ex_head, column=idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        for idx, column in enumerate(columns, start=1):
            cell = info.cell(row=ex_head + 1, column=idx, value=example[column])
            cell.border = border
            cell.alignment = wrap

        info.column_dimensions["A"].width = 18
        info.column_dimensions["B"].width = 14
        info.column_dimensions["C"].width = 70

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def _row_to_listing_attributes(row: pd.Series) -> Dict[str, List[Dict[str, Any]]]:
    """Translate a spreadsheet row into SP-API attribute patches."""
    attributes: Dict[str, List[Dict[str, Any]]] = {}
    bullet_values: List[str] = []

    for column, attribute in LISTING_TEXT_FIELDS.items():
        value = row.get(column)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        text = str(value).strip()
        if not text:
            continue

        if attribute == "bullet_point":
            bullet_values.append(text)
            continue

        if attribute == "generic_keyword":
            attributes[attribute] = [{"value": text}]
        else:
            attributes[attribute] = [{"value": text}]

    if bullet_values:
        attributes["bullet_point"] = [{"value": v} for v in bullet_values]

    return attributes


def _row_listing_snapshot(row: pd.Series) -> Dict[str, Any]:
    fields: Dict[str, Any] = {}
    for column in LISTING_TEXT_FIELDS:
        value = row.get(column)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        fields[column] = str(value)
    return fields


def _cell(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def parse_import_rows(
    file_bytes: bytes, filename: str
) -> tuple[List[Dict[str, str]], List[BulkRowError]]:
    """Parse a CSV/Excel catalog import into clean rows + per-row errors.

    Pure function: no DB access. Headers are matched case-insensitively after
    trimming. ``asin`` is required and validated against ASIN_PATTERN; duplicate
    ASINs within the file keep only the first occurrence.
    """
    name = (filename or "").lower()
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(
                io.BytesIO(file_bytes), dtype=str, sep=None, engine="python", encoding="utf-8-sig"
            )
        elif name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        else:
            raise CatalogOperationError("File must be a .csv, .xlsx or .xls file")
    except CatalogOperationError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise CatalogOperationError(f"Could not parse file: {exc}") from exc

    df.columns = [str(c).strip().lstrip("﻿").strip().lower() for c in df.columns]
    if "asin" not in df.columns:
        raise CatalogOperationError("File must contain an 'asin' column")

    rows: List[Dict[str, str]] = []
    errors: List[BulkRowError] = []
    seen: set[str] = set()

    for row_idx, row in df.iterrows():
        row_number = int(row_idx) + 2  # +1 for 0-based, +1 for header
        asin = _cell(row.get("asin"))
        if not _ASIN_RE.match(asin):
            errors.append(
                BulkRowError(
                    row=row_number,
                    asin=asin or None,
                    error="ASIN is required and must be 10 uppercase letters or digits",
                    code=BulkErrorCode.INVALID_INPUT,
                )
            )
            continue
        if asin in seen:
            continue
        seen.add(asin)

        rows.append(
            {
                "asin": asin,
                "sku": _cell(row.get("sku")),
                "title": _cell(row.get("title")),
                "brand": _cell(row.get("brand")),
                "category": _cell(row.get("category")),
            }
        )

    return rows, errors


def import_template_bytes() -> bytes:
    """Return a tiny CSV template (header + one example row) for manual import."""
    header = ",".join(IMPORT_COLUMNS)
    example = "B08N5WRWNW,EXAMPLE-SKU-001,Esempio prodotto,Marca,Categoria"
    return ("﻿" + header + "\n" + example + "\n").encode("utf-8")


_CURRENCY_BY_MARKETPLACE = {
    "IT": "EUR", "DE": "EUR", "FR": "EUR", "ES": "EUR", "NL": "EUR",
    "BE": "EUR", "PL": "PLN", "SE": "SEK", "TR": "TRY",
    "UK": "GBP", "GB": "GBP",
    "US": "USD", "CA": "CAD", "MX": "MXN", "BR": "BRL",
    "JP": "JPY", "AU": "AUD", "IN": "INR", "AE": "AED", "SG": "SGD",
}


def _marketplace_currency(country_code: str) -> str:
    return _CURRENCY_BY_MARKETPLACE.get((country_code or "").upper(), "EUR")
