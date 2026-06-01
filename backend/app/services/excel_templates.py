"""Excel template styles and renderer for professional export formatting."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_FONT_NAME = "Calibri"
DEFAULT_FONT_SIZE = 10


@dataclass
class TemplateStyle:
    """Visual parameters for an Excel template."""

    name: str
    # Header
    header_fill_color: str
    header_font_color: str
    header_font_size: int = 11
    # Rows
    row_even_fill: str = "FFFFFF"
    row_odd_fill: str = "F9F9F9"
    # KPI / accent
    kpi_accent_color: str = "333333"
    kpi_accent_size: int = 14
    kpi_accent_fill: str | None = None
    # Borders
    border_color: str = "E0E0E0"
    # Title banner
    title_fill_color: str | None = None
    title_font_color: str = "FFFFFF"
    title_font_size: int = 16
    subtitle_font_size: int = 11
    # Accent used for tab color, separators and totals row
    accent_color: str = "1F4E79"
    totals_fill: str = "DCE6F1"


CLEAN = TemplateStyle(
    name="clean",
    header_fill_color="2F5496",
    header_font_color="FFFFFF",
    row_even_fill="FFFFFF",
    row_odd_fill="F2F5FA",
    kpi_accent_color="2F5496",
    kpi_accent_size=14,
    border_color="D9DEE6",
    title_fill_color="2F5496",
    title_font_color="FFFFFF",
    title_font_size=16,
    accent_color="2F5496",
    totals_fill="E8EDF5",
)

CORPORATE = TemplateStyle(
    name="corporate",
    header_fill_color="1F4E79",
    header_font_color="FFFFFF",
    row_even_fill="FFFFFF",
    row_odd_fill="E4ECF5",
    kpi_accent_color="1F4E79",
    kpi_accent_size=14,
    border_color="C2D2E5",
    title_fill_color="1F4E79",
    title_font_color="FFFFFF",
    title_font_size=16,
    accent_color="1F4E79",
    totals_fill="D2E0F0",
)

EXECUTIVE = TemplateStyle(
    name="executive",
    header_fill_color="1B2631",
    header_font_color="FFFFFF",
    row_even_fill="FFFFFF",
    row_odd_fill="EEF1F4",
    kpi_accent_color="C9962E",
    kpi_accent_size=14,
    kpi_accent_fill="C9962E",
    border_color="C3CAD2",
    title_fill_color="1B2631",
    title_font_color="FFFFFF",
    title_font_size=16,
    accent_color="C9962E",
    totals_fill="E2E6EA",
)

TEMPLATES: dict[str, TemplateStyle] = {
    "clean": CLEAN,
    "corporate": CORPORATE,
    "executive": EXECUTIVE,
}


class ExcelTemplateRenderer:
    """Renders styled Excel sheets using a TemplateStyle."""

    def __init__(self, style: TemplateStyle) -> None:
        self.style = style
        self._header_fill = PatternFill(
            start_color=style.header_fill_color,
            end_color=style.header_fill_color,
            fill_type="solid",
        )
        self._header_font = Font(
            name=DEFAULT_FONT_NAME,
            bold=True,
            color=style.header_font_color,
            size=style.header_font_size,
        )
        self._body_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
        self._even_fill = PatternFill(
            start_color=style.row_even_fill,
            end_color=style.row_even_fill,
            fill_type="solid",
        )
        self._odd_fill = PatternFill(
            start_color=style.row_odd_fill,
            end_color=style.row_odd_fill,
            fill_type="solid",
        )
        self._border = Border(
            left=Side(style="thin", color=style.border_color),
            right=Side(style="thin", color=style.border_color),
            top=Side(style="thin", color=style.border_color),
            bottom=Side(style="thin", color=style.border_color),
        )
        self._totals_fill = PatternFill(
            start_color=style.totals_fill,
            end_color=style.totals_fill,
            fill_type="solid",
        )
        self._totals_top = Side(style="medium", color=style.accent_color)
        self._kpi_font = Font(
            name=DEFAULT_FONT_NAME,
            bold=True,
            color=style.kpi_accent_color,
            size=style.kpi_accent_size,
        )
        self._kpi_fill = (
            PatternFill(
                start_color=style.kpi_accent_fill,
                end_color=style.kpi_accent_fill,
                fill_type="solid",
            )
            if style.kpi_accent_fill
            else None
        )
        self._title_fill = (
            PatternFill(
                start_color=style.title_fill_color,
                end_color=style.title_fill_color,
                fill_type="solid",
            )
            if style.title_fill_color
            else None
        )
        self._title_font = Font(
            name=DEFAULT_FONT_NAME,
            bold=True,
            color=style.title_font_color,
            size=style.title_font_size,
        )
        self._subtitle_font = Font(
            name=DEFAULT_FONT_NAME,
            color=style.title_font_color,
            size=style.subtitle_font_size,
        )

    def write_title_banner(
        self,
        ws: Worksheet,
        title: str,
        subtitle: str,
        metadata_rows: list[tuple[str, str]],
    ) -> int:
        """Write a branded title banner area. Returns the next free row."""
        num_cols = max(6, 2 + len(metadata_rows))
        self._apply_tab_color(ws)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = self._title_font
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if self._title_fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=1, column=col).fill = self._title_fill
            cell.fill = self._title_fill
        ws.row_dimensions[1].height = 42

        # Subtitle row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        cell = ws.cell(row=2, column=1, value=subtitle)
        cell.font = self._subtitle_font
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if self._title_fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=2, column=col).fill = self._title_fill
            cell.fill = self._title_fill
        ws.row_dimensions[2].height = 22

        # Metadata rows
        label_font = Font(name=DEFAULT_FONT_NAME, bold=True, size=10, color="595959")
        value_font = Font(name=DEFAULT_FONT_NAME, size=10, color="262626")
        row_num = 4
        for label, value in metadata_rows:
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.cell(row=row_num, column=2, value=value).font = value_font
            row_num += 1

        # Thin accent separator under the metadata block
        separator = Border(bottom=Side(style="thin", color=self.style.accent_color))
        for col in range(1, num_cols + 1):
            ws.cell(row=row_num, column=col).border = separator
        ws.row_dimensions[row_num].height = 6

        return row_num + 2

    def write_data_sheet(
        self,
        ws: Worksheet,
        rows: list[dict[str, Any]],
        columns: list[str],
        headers: list[str],
        start_row: int = 1,
        number_formats: dict[str, str] | None = None,
        totals: bool = False,
        total_columns: list[str] | None = None,
    ) -> None:
        """Write a data table with styled headers and alternating rows.

        ``number_formats`` maps a column key to an openpyxl number format string
        (e.g. ``'€#,##0.00'`` or ``'0.0%'``). Columns without an entry fall back
        to integer/float defaults. Set ``totals`` to append a summed totals row;
        ``total_columns`` restricts which numeric columns are summed (defaults to
        all numeric columns when ``totals`` is on).
        """
        number_formats = number_formats or {}
        self._write_header(ws, columns, headers, start_row)

        sums: dict[str, float] = {}
        last_data_row = start_row
        for row_idx, row_data in enumerate(rows):
            excel_row = start_row + 1 + row_idx
            last_data_row = excel_row
            fill = self._even_fill if row_idx % 2 == 0 else self._odd_fill
            for col_idx, col_key in enumerate(columns, 1):
                value = self._normalize_value(row_data.get(col_key, ""))
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = self._body_font
                cell.fill = fill
                cell.border = self._border
                cell.alignment = Alignment(vertical="center")
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = self._format_for(col_key, value, number_formats)
                    sums[col_key] = sums.get(col_key, 0.0) + float(value)

        if totals and rows:
            self._write_totals_row(
                ws,
                columns,
                last_data_row + 1,
                sums,
                total_columns,
                number_formats,
            )

        self._auto_width(ws, columns, headers, start_row, len(rows))
        self._finalize_sheet(ws, columns, headers, start_row, len(rows))

    def write_summary_sheet(
        self,
        ws: Worksheet,
        summary_rows: list[dict[str, Any]],
        columns: list[str],
        headers: list[str],
        start_row: int = 1,
        number_formats: dict[str, str] | None = None,
    ) -> None:
        """Write a summary/KPI table with accent highlighting on values."""
        number_formats = number_formats or {}
        self._write_header(ws, columns, headers, start_row)

        for row_idx, row_data in enumerate(summary_rows):
            excel_row = start_row + 1 + row_idx
            fill = self._even_fill if row_idx % 2 == 0 else self._odd_fill
            for col_idx, col_key in enumerate(columns, 1):
                value = self._normalize_value(row_data.get(col_key, ""))
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = self._body_font
                cell.fill = fill
                cell.border = self._border
                cell.alignment = Alignment(vertical="center")

                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = self._format_for(col_key, value, number_formats)

                # KPI accent on current_value column
                if col_key == "current_value":
                    cell.font = self._kpi_font
                    if self._kpi_fill:
                        cell.fill = self._kpi_fill
                        cell.font = Font(
                            name=DEFAULT_FONT_NAME,
                            bold=True,
                            color="FFFFFF",
                            size=self.style.kpi_accent_size,
                        )

                # Change % formatting (green/red)
                if col_key == "change_percent" and value != "":
                    self.apply_change_formatting(cell, value)

        self._auto_width(ws, columns, headers, start_row, len(summary_rows))
        self._finalize_sheet(ws, columns, headers, start_row, len(summary_rows))

    def apply_change_formatting(self, cell: Any, value: Any) -> None:
        """Apply green for positive, red for negative change values."""
        try:
            num = float(value)
        except (TypeError, ValueError):
            return
        if num > 0:
            cell.font = Font(name=DEFAULT_FONT_NAME, bold=True, color="27AE60")
        elif num < 0:
            cell.font = Font(name=DEFAULT_FONT_NAME, bold=True, color="E74C3C")

    def _write_header(
        self,
        ws: Worksheet,
        columns: list[str],
        headers: list[str],
        start_row: int,
    ) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = self._border
        ws.row_dimensions[start_row].height = 30

    def _write_totals_row(
        self,
        ws: Worksheet,
        columns: list[str],
        excel_row: int,
        sums: dict[str, float],
        total_columns: list[str] | None,
        number_formats: dict[str, str],
    ) -> None:
        label = "Totale" if self.style.name in ("corporate", "executive") else "Total"
        bold = Font(name=DEFAULT_FONT_NAME, bold=True, size=DEFAULT_FONT_SIZE)
        targets = set(total_columns) if total_columns is not None else set(sums.keys())
        for col_idx, col_key in enumerate(columns, 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = self._totals_fill
            cell.font = bold
            cell.border = Border(
                top=self._totals_top,
                left=Side(style="thin", color=self.style.border_color),
                right=Side(style="thin", color=self.style.border_color),
                bottom=Side(style="thin", color=self.style.border_color),
            )
            cell.alignment = Alignment(vertical="center")
            if col_idx == 1:
                cell.value = label
            elif col_key in targets and col_key in sums:
                cell.value = sums[col_key]
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = self._format_for(
                    col_key, sums[col_key], number_formats
                )

    def _format_for(
        self,
        col_key: str,
        value: Any,
        number_formats: dict[str, str],
    ) -> str:
        if col_key in number_formats:
            return number_formats[col_key]
        if isinstance(value, float):
            return "#,##0.00"
        return "#,##0"

    def _apply_tab_color(self, ws: Worksheet) -> None:
        ws.sheet_properties.tabColor = self.style.accent_color

    def _finalize_sheet(
        self,
        ws: Worksheet,
        columns: list[str],
        headers: list[str],
        header_row: int,
        num_rows: int,
    ) -> None:
        """Apply freeze panes, auto-filter and print setup for a finished table."""
        self._apply_tab_color(ws)
        last_col = get_column_letter(len(columns))

        # Freeze everything above and including the header row.
        ws.freeze_panes = f"A{header_row + 1}"

        # Auto-filter across the header + data range.
        last_row = header_row + max(num_rows, 0)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

        # Print setup: landscape, fit to one page wide, repeat header on every page.
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f"{header_row}:{header_row}"

    def _normalize_value(self, value: Any) -> Any:
        """Convert values to Excel-friendly types."""
        if isinstance(value, bool):
            return value
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return ""
        return value

    def _auto_width(
        self,
        ws: Worksheet,
        columns: list[str],
        headers: list[str],
        start_row: int,
        num_rows: int,
    ) -> None:
        """Auto-fit column widths based on header and data content."""
        for col_idx, header in enumerate(headers, 1):
            # Headers wrap, so size against the longest word rather than full text.
            max_len = max((len(w) for w in str(header).split()), default=0)
            # Sample first 50 rows for width
            for row_offset in range(min(num_rows, 50)):
                cell = ws.cell(row=start_row + 1 + row_offset, column=col_idx)
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
